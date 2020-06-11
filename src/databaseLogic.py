from openpyxl import Workbook
from openpyxl import load_workbook 


wb=load_workbook('./db/shopdata.xlsx')

def getCustomerID(invoiceID):
    invoice_ws = wb["invoice"]
    for row in invoice_ws['A']:
        if row.value == invoiceID:
            customerID = invoice_ws.cell(row=row.row, column=row.column+1).value 
    return customerID

def getCustomerName(customerID):
    customer_ws = wb['customer']    
    for row in customer_ws['A']:
        if row.value == customerID:
            customerFName = customer_ws.cell(row=row.row, column=row.column+1).value
            customerLName = customer_ws.cell(row=row.row, column=row.column+2).value
    return f"{customerFName} {customerLName}"

def getItems(invoiceID):
    items_ws=wb['invoice_list']
    product_ws = wb['product']
    itemsList=[]
    for row in items_ws['B']:
        if row.value == invoiceID:
            a={}
            a['item_id']=items_ws.cell(row=row.row, column=row.column-1).value
            a['qty']=items_ws.cell(row=row.row, column=row.column+1).value
            a['price'] =0
            itemsList.append(a)
    for row in product_ws['A']:
        for eachItem in itemsList:
            if eachItem['item_id'] == row.value:
                eachItem['name']= product_ws.cell(row=row.row, column=row.column+1).value
                eachItem['price']= product_ws.cell(row=row.row, column=row.column+2).value
    return itemsList


def create_invoice(invoiceID):
    customerID =getCustomerID(invoiceID) 
    full_name =getCustomerName(customerID)
    items= getItems(invoiceID)
    return {
        'invoiceID':invoiceID,
        'name':full_name,
        'items': items
        }
    