import pytest
import os
from openpyxl import load_workbook
from databaseLogic import create_invoice, getCustomerID

wb=load_workbook('./db/shopdata.xlsx')
def test_answer():    
    if "shopdata.xlsx" in os.listdir('./db'):
        x = True
    assert x == True

def test_ws():
    sheets =wb.sheetnames 
    x = 'customer' in sheets 
    assert x == True
    x = 'invoice' in sheets 
    assert x == True
    x = 'invoice_list' in sheets 
    assert x == True
    x = 'product' in sheets 
    assert x == True

def test_invoice():
   assert type(create_invoice(1)) == dict
   assert getCustomerID(1) == 5
   assert create_invoice(1)['name'] == 'Chloe Jones'
